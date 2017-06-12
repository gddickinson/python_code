#-------------------------------------------------------------------------------
# Name:        Step_6_Interact_with_Shade
# Purpose:      Run Shadalator from ArcGIS - import/export data to excel file
#
# Author:      George
#
# Created:     08/05/2017
# Copyright:   (c) George 2017
# Licence:     <your licence>
#-------------------------------------------------------------------------------

# Import system modules
from __future__ import division, print_function
import sys
import os
import gc
import time
import traceback
from datetime import timedelta
from math import radians, sin, cos, ceil, sqrt
from collections import defaultdict, OrderedDict
import numpy as np
import arcpy
from arcpy import env

import openpyxl
import win32com.client
import io

#enable garbage collection
gc.enable()

################################################################################
##nodes_fc = r"C:\Google Drive\SiCr_Digitization\shadeOutput\nodes_fc.shp"
##exel_filename = r"C:\Google Drive\SiCr_Digitization\scripts\blankMainMenu.xlsx"
##exel_filename2 = r"C:\Google Drive\SiCr_Digitization\scripts\blank.xlsx"
##runName = "SilverCreek"
##saveName = r"C:\Google Drive\SiCr_Digitization\scripts\output.xlsx"
##saveName2 = r"C:\Google Drive\SiCr_Digitization\scripts\output2.xlsx"
##sheetName = "Main Menu"
##excel_shade = openpyxl.load_workbook(exel_filename)
##excel_output = openpyxl.load_workbook(exel_filename2)
##startDate = "08/01/2016"
##numberDays = 1
##lat = 43.28
##longi = -114.02
##timeZone = -8
##daylightSavings = "Yes"
##globalRiparianZoneWidth = 30.48
##shadeCalculationMethod = "Chen"
##vegCodes = "On"
##cloudCover = 0
##brasVisibilityFactor = 2
##ryanStolzenbachFactor = 0.8
##elevationOrZone = "Elevation"
##riparianExtinction = "On"


################################################################################

nodes_fc = arcpy.GetParameterAsText(0)
exel_filename = r"C:\Google Drive\SiCr_Digitization\scripts\blankMainMenu.xlsx"
exel_filename2 = r"C:\Google Drive\SiCr_Digitization\scripts\blank.xlsx"
runName = arcpy.GetParameterAsText(1)
saveName = r"C:\Google Drive\SiCr_Digitization\scripts\output.xlsx"
saveName2 = r"C:\Google Drive\SiCr_Digitization\scripts\output2.xlsx"
sheetName = "Main Menu"


with open(exel_filename, "rb") as f:
    in_mem_file1 = io.BytesIO(f.read())


with open(exel_filename2, "rb") as f:
    in_mem_file2 = io.BytesIO(f.read())

excel_shade = openpyxl.load_workbook(in_mem_file1)
excel_output = openpyxl.load_workbook(in_mem_file2)

startDate = arcpy.GetParameterAsText(2)
numberDays = arcpy.GetParameterAsText(3)
lat = arcpy.GetParameterAsText(4)
longi = arcpy.GetParameterAsText(5)

timeZone = arcpy.GetParameterAsText(6)
daylightSavings = arcpy.GetParameterAsText(7)
globalRiparianZoneWidth = arcpy.GetParameterAsText(8)
shadeCalculationMethod = arcpy.GetParameterAsText(9)
vegCodes = arcpy.GetParameterAsText(10)
cloudCover = arcpy.GetParameterAsText(11)
brasVisibilityFactor = arcpy.GetParameterAsText(12)
ryanStolzenbachFactor = arcpy.GetParameterAsText(13)
elevationOrZone = arcpy.GetParameterAsText(14)
riparianExtinction = arcpy.GetParameterAsText(15)


################################################################################

shadelatorPath = r"C:\Google Drive\SiCr_Digitization\scripts\shade_GD.xlsm"
shadelatorSavePath = r"C:\Google Drive\SiCr_Digitization\scripts\shadeResults_GD.xlsm"


def extractData(fc,fieldList):
    ans = []
    cursor = arcpy.da.SearchCursor(fc, fieldList)
    for row in cursor:
        ans.append(row[0])
    return ans


def writeColumn(data, workbook, worksheet, startRow, startColumn):
    cell = startColumn+str(startRow)
    activeSheet = workbook[worksheet]
    for i in data:
        activeSheet[cell] = i
        startRow += 1
        cell = startColumn+str(startRow)

def writeCell(data, workbook, worksheet, row, column):
    cell = column+str(row)
    activeSheet = workbook[worksheet]
    activeSheet[cell] = data

def run_macro(pathName):
    print('macro-running')
    print(pathName)

    macro1 = "Sheet1.CommandButton1_Click"
    macro2 = "Procedures"
    macro3 = "Sheet1.CommandButton2_Click"

    #this part runs the macro from excel
    if os.path.exists(pathName):
        xl=win32com.client.Dispatch("Excel.Application")
        #xl.Visible = 1
        xl.Workbooks.Open(Filename=pathName, ReadOnly=1)
        xl.DisplayAlerts = False

        xl.Application.Run(macro1)
        print("Data Added To Shade-a-lator")
        arcpy.AddMessage("Data Added To Shade-a-lator")


        xl.Application.Run(macro2)


        print("Export Data from Shade-a-lator")
        arcpy.AddMessage("Export Data from Shade-a-lator")

        xl.Application.Run(macro3)
        print("Data exported from Shade-a-lator")
        arcpy.AddMessage("Data exported from Shade-a-lator!")

        #xl.Save(shadelatorSavePath)
        map(lambda book: book.Close(False), xl.Workbooks)
        xl.Application.Quit() # Comment this out if excel script closes
        del xl

    print('File refreshed!')


################################################################################



try:
    arcpy.AddMessage("Step 6: Exporting Data to Excel")
    print("Step 6: Exporting Data to Excel")

    #keeping track of time
    startTime= time.time()

    # Check if the node fc exists
    if not arcpy.Exists(nodes_fc):
        arcpy.AddError("\nThis output does not exist: \n" +
                       "{0}\n".format(nodes_fc))
        sys.exit("This output does not exist: \n" +
                 "{0}\n".format(nodes_fc))


    # Get a list of existing fields
    existingFields = [f.name for f in arcpy.ListFields(nodes_fc)]

    #Make arrays storing field data

    STREAM_ID = extractData(nodes_fc, 'STREAM_ID')
    FID = extractData(nodes_fc, 'FID')
    NODEID = extractData(nodes_fc, 'NODE_ID')
    LONGITUDE = extractData(nodes_fc, 'LONGITUDE')
    LATITUDE = extractData(nodes_fc, 'LATITUDE')
    STRM_AZMTH = extractData(nodes_fc, 'STRM_AZMTH')
    CHANWIDTH = extractData(nodes_fc, 'CHANWIDTH')
    LEFT = extractData(nodes_fc, 'LEFT')
    RIGHT = extractData(nodes_fc, 'RIGHT')


    def setToZero(listName):
        ans =[]
        for row in listName:
            row = float(row)
            if row < 0:
                row = 0
            ans.append(row)
        return ans

    #setting elevation less than 0 to 0 to avoid error in shadelator
    ELEVATION = setToZero(extractData(nodes_fc, 'ELEVATION'))

    #setting topographic angles less than 0 to 0 to avoid error in shadelator
    TOPO_W = setToZero(extractData(nodes_fc, 'TOPO_W'))
    TOPO_S = setToZero(extractData(nodes_fc, 'TOPO_S'))
    TOPO_E = setToZero(extractData(nodes_fc, 'TOPO_E'))

    #adding offest to get round problem with Riparian Code 0 generation in shadelator
    def offsetBy1(listName):
        ans =[]
        for row in listName:
            row = int(row)
            if row < 0:
                row = 0
            ans.append(row+1)
        return ans

    LC_T1_S1 = offsetBy1(extractData(nodes_fc, 'LC_T1_S1'))
    LC_T1_S2 = offsetBy1(extractData(nodes_fc, 'LC_T1_S2'))
    LC_T1_S3 = offsetBy1(extractData(nodes_fc, 'LC_T1_S3'))
    LC_T1_S4 = offsetBy1(extractData(nodes_fc, 'LC_T1_S4'))
    LC_T1_S5 = offsetBy1(extractData(nodes_fc, 'LC_T1_S5'))
    LC_T1_S6 = offsetBy1(extractData(nodes_fc, 'LC_T1_S6'))
    LC_T1_S7 = offsetBy1(extractData(nodes_fc, 'LC_T1_S7'))
    LC_T1_S8 = offsetBy1(extractData(nodes_fc, 'LC_T1_S8'))
    LC_T1_S9 = offsetBy1(extractData(nodes_fc, 'LC_T1_S9'))

    LC_T2_S1 = offsetBy1(extractData(nodes_fc, 'LC_T2_S1'))
    LC_T2_S2 = offsetBy1(extractData(nodes_fc, 'LC_T2_S2'))
    LC_T2_S3 = offsetBy1(extractData(nodes_fc, 'LC_T2_S3'))
    LC_T2_S4 = offsetBy1(extractData(nodes_fc, 'LC_T2_S4'))
    LC_T2_S5 = offsetBy1(extractData(nodes_fc, 'LC_T2_S5'))
    LC_T2_S6 = offsetBy1(extractData(nodes_fc, 'LC_T2_S6'))
    LC_T2_S7 = offsetBy1(extractData(nodes_fc, 'LC_T2_S7'))
    LC_T2_S8 = offsetBy1(extractData(nodes_fc, 'LC_T2_S8'))
    LC_T2_S9 = offsetBy1(extractData(nodes_fc, 'LC_T2_S9'))

    ELE_T1_S1 = extractData(nodes_fc, 'ELE_T1_S1')
    ELE_T1_S2 = extractData(nodes_fc, 'ELE_T1_S2')
    ELE_T1_S3 = extractData(nodes_fc, 'ELE_T1_S3')
    ELE_T1_S4 = extractData(nodes_fc, 'ELE_T1_S4')
    ELE_T1_S5 = extractData(nodes_fc, 'ELE_T1_S5')
    ELE_T1_S6 = extractData(nodes_fc, 'ELE_T1_S6')
    ELE_T1_S7 = extractData(nodes_fc, 'ELE_T1_S7')
    ELE_T1_S8 = extractData(nodes_fc, 'ELE_T1_S8')
    ELE_T1_S9 = extractData(nodes_fc, 'ELE_T1_S9')

    ELE_T2_S1 = extractData(nodes_fc, 'ELE_T2_S1')
    ELE_T2_S2 = extractData(nodes_fc, 'ELE_T2_S2')
    ELE_T2_S3 = extractData(nodes_fc, 'ELE_T2_S3')
    ELE_T2_S4 = extractData(nodes_fc, 'ELE_T2_S4')
    ELE_T2_S5 = extractData(nodes_fc, 'ELE_T2_S5')
    ELE_T2_S6 = extractData(nodes_fc, 'ELE_T2_S6')
    ELE_T2_S7 = extractData(nodes_fc, 'ELE_T2_S7')
    ELE_T2_S8 = extractData(nodes_fc, 'ELE_T2_S8')
    ELE_T2_S9 = extractData(nodes_fc, 'ELE_T2_S9')


    numberNodes = len(FID)
    DISTANCE = []
    CHANNEL_INC = []

    dist = 0
    incision = 1.0
    for i in range(numberNodes):
        DISTANCE.append(dist)
        CHANNEL_INC.append(incision)
        dist += 50

    #write data to excel file

    writeColumn(STREAM_ID,excel_output,"Sheet1",2,'A')
    writeColumn(FID,excel_output,"Sheet1",2,'B')
    writeColumn(ELEVATION,excel_output,"Sheet1",2,'C')
    writeColumn(LATITUDE,excel_output,"Sheet1",2,'D')
    writeColumn(LONGITUDE,excel_output,"Sheet1",2,'E')


    writeColumn(STREAM_ID,excel_shade,sheetName,17,'A')
    writeColumn(ELEVATION,excel_shade,sheetName,17,'C')
    writeColumn(STRM_AZMTH,excel_shade,sheetName,17,'E')
    writeColumn(CHANWIDTH,excel_shade,sheetName,17,'H')
    writeColumn(CHANWIDTH,excel_shade,sheetName,17,'I') #both width columns kept the same
    writeColumn(TOPO_W,excel_shade,sheetName,17,'L')
    writeColumn(TOPO_S,excel_shade,sheetName,17,'M')
    writeColumn(TOPO_E,excel_shade,sheetName,17,'N')
    writeColumn(DISTANCE,excel_shade,sheetName,17,'B')
    writeColumn(CHANNEL_INC,excel_shade,sheetName,17,'K')

    writeColumn(LC_T1_S1,excel_shade,sheetName,17,'P')
    writeColumn(LC_T1_S2,excel_shade,sheetName,17,'Q')
    writeColumn(LC_T1_S3,excel_shade,sheetName,17,'R')
    writeColumn(LC_T1_S4,excel_shade,sheetName,17,'S')
    writeColumn(LC_T1_S5,excel_shade,sheetName,17,'T')
    writeColumn(LC_T1_S6,excel_shade,sheetName,17,'U')
    writeColumn(LC_T1_S7,excel_shade,sheetName,17,'V')
    writeColumn(LC_T1_S8,excel_shade,sheetName,17,'W')
    writeColumn(LC_T1_S9,excel_shade,sheetName,17,'X')

    writeColumn(LC_T2_S1,excel_shade,sheetName,17,'Y')
    writeColumn(LC_T2_S2,excel_shade,sheetName,17,'Z')
    writeColumn(LC_T2_S3,excel_shade,sheetName,17,'AA')
    writeColumn(LC_T2_S4,excel_shade,sheetName,17,'AB')
    writeColumn(LC_T2_S5,excel_shade,sheetName,17,'AC')
    writeColumn(LC_T2_S6,excel_shade,sheetName,17,'AD')
    writeColumn(LC_T2_S7,excel_shade,sheetName,17,'AE')
    writeColumn(LC_T2_S8,excel_shade,sheetName,17,'AF')
    writeColumn(LC_T2_S9,excel_shade,sheetName,17,'AG')

    writeColumn(ELE_T1_S1,excel_shade,sheetName,17,'AH')
    writeColumn(ELE_T1_S2,excel_shade,sheetName,17,'AI')
    writeColumn(ELE_T1_S3,excel_shade,sheetName,17,'AJ')
    writeColumn(ELE_T1_S4,excel_shade,sheetName,17,'AK')
    writeColumn(ELE_T1_S5,excel_shade,sheetName,17,'AL')
    writeColumn(ELE_T1_S6,excel_shade,sheetName,17,'AM')
    writeColumn(ELE_T1_S7,excel_shade,sheetName,17,'AN')
    writeColumn(ELE_T1_S8,excel_shade,sheetName,17,'AO')
    writeColumn(ELE_T1_S9,excel_shade,sheetName,17,'AP')

    writeColumn(ELE_T2_S1,excel_shade,sheetName,17,'AQ')
    writeColumn(ELE_T2_S2,excel_shade,sheetName,17,'AR')
    writeColumn(ELE_T2_S3,excel_shade,sheetName,17,'AS')
    writeColumn(ELE_T2_S4,excel_shade,sheetName,17,'AT')
    writeColumn(ELE_T2_S5,excel_shade,sheetName,17,'AU')
    writeColumn(ELE_T2_S6,excel_shade,sheetName,17,'AV')
    writeColumn(ELE_T2_S7,excel_shade,sheetName,17,'AW')
    writeColumn(ELE_T2_S8,excel_shade,sheetName,17,'AX')
    writeColumn(ELE_T2_S9,excel_shade,sheetName,17,'AY')


    writeCell(startDate,excel_shade,sheetName,4,'D')
    writeCell(numberDays,excel_shade,sheetName,5,'D')
    writeCell(lat,excel_shade,sheetName,8,'D')
    writeCell(longi,excel_shade,sheetName,9,'D')
    writeCell(runName,excel_shade,sheetName,2,'D')

    writeCell(timeZone,excel_shade,sheetName,6,'D')
    writeCell(daylightSavings,excel_shade,sheetName,7,'D')
    writeCell(globalRiparianZoneWidth,excel_shade,sheetName,10,'D')
    writeCell(shadeCalculationMethod,excel_shade,sheetName,11,'D')
    writeCell(vegCodes,excel_shade,sheetName,3,'I')
    writeCell(cloudCover,excel_shade,sheetName,5,'I')
    writeCell(brasVisibilityFactor,excel_shade,sheetName,6,'I')
    writeCell(ryanStolzenbachFactor,excel_shade,sheetName,7,'I')
    writeCell(elevationOrZone,excel_shade,sheetName,8,'I')
    writeCell(riparianExtinction,excel_shade,sheetName,9,'I')

    excel_shade.save(saveName)
    excel_shade.close()
    excel_shade = None

    excel_output.save(saveName2)
    excel_output.close()
    excel_output= None

    print("Data Exported To Excel")
    arcpy.AddMessage("Data Exported To Excel")

    print("--------------------------------------")
    arcpy.AddMessage("--------------------------------------")

    print("Step 7: Run Shade-a-lator")
    arcpy.AddMessage("Step 7: Run Shade-a-lator")

    print("Run Macros")
    arcpy.AddMessage("Run Macros")

    run_macro(shadelatorPath)



    print("Shade-a-lator finished")
    arcpy.AddMessage("Shade-a-lator finished")

    endTime = time.time()

    elapsedmin= ceil(((endTime - startTime) / 60)* 10)/10

    print("Process Complete in {0} minutes".format(elapsedmin))
    arcpy.AddMessage("Process Complete in %s minutes" % (elapsedmin))


# For arctool errors
except arcpy.ExecuteError:
    msgs = arcpy.GetMessages(2)
    arcpy.AddError(msgs)
    print(msgs)

# For other errors
except:
    tbinfo = traceback.format_exc()

    pymsg = "PYTHON ERRORS:\n" + tbinfo + "\nError Info:\n" +str(sys.exc_info()[1])
    msgs = "ArcPy ERRORS:\n" + arcpy.GetMessages(2) + "\n"

    arcpy.AddError(pymsg)
    arcpy.AddError(msgs)

    print(pymsg)
    print(msgs)
